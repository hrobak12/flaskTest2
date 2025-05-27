import os, secrets
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, abort, jsonify, Response, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy import func, and_, asc, desc, extract, or_

from datetime import datetime
import bcrypt
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, mm #, A5, landscape, portrait
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from barcode import Code128
from barcode.writer import ImageWriter
from transliterate import translit

from models import (db, User, RefillDept, PrinterModel, CustomerEquipment, Cartridges, CartridgeStatus, EventLog,
                    CartridgeModel, CompatibleCartridges, Contracts, ContractsServicesBalance, CompatibleServices)

from config import status_map
from services import *

app = Flask(__name__)
# Тільки для розробки
# Виведе випадковий 64-символьний ключ, якщо інший не заданий в системі
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cartridge.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Допоміжні функції
def hash_password(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def check_password(password, hashed):
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def admin_required(f):
    def wrapper(*args, **kwargs):
        if current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

#*************
#тут може бути скрипт міграції, якщо потрібно

with app.app_context():
    db.create_all()

#*************


@app.route('/')
@app.route('/index')
@login_required
def index():
    return render_template('index.html', user=current_user, RefillDept=RefillDept)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password(password, user.password):
            login_user(user)
            return redirect(url_for('index'))
        flash('Невірний логін або пароль')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# Відображення подій обробки картриджів та керування ними
# Основний маршрут для відображення подій
@app.route('/cartridge_status')
@login_required
@admin_required
def cartridge_status():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # Базовий запит із JOIN для оптимізації
    query = db.session.query(CartridgeStatus, Cartridges.serial_num, RefillDept.deptname)\
                      .outerjoin(Cartridges, Cartridges.id == CartridgeStatus.cartridge_id)\
                      .outerjoin(RefillDept, RefillDept.id == CartridgeStatus.exec_dept)\
                      .order_by(CartridgeStatus.date_ofchange.desc())

    if search:
        query = query.filter(Cartridges.serial_num.ilike(f'%{search}%'))

    # Пагінація
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    statuses = [(s[0], s[1] or 'Не вказано', s[2] or 'Не вказано') for s in pagination.items]

    return render_template('cartridge_status.html',
                           statuses=statuses,
                           search=search,
                           pagination=pagination)

@app.route('/update_status/<int:status_id>', methods=['POST'])
@login_required
@admin_required
def update_status(status_id):
    status = CartridgeStatus.query.get_or_404(status_id)
    new_status = int(request.form['status'])
    status.status = new_status
    status.user_updated = current_user.id
    status.time_updated = datetime.now()  # Оновлюємо час зміни
    event = EventLog(
        table_name='cartrg_status',
        event_type=2,  # Оновлення статусу
        user_updated=current_user.id
    )
    db.session.add(event)
    db.session.commit()
    flash('Статус оновлено!')
    return redirect(url_for('cartridge_status'))

@app.route('/delete_status/<int:status_id>', methods=['POST'])
@login_required
@admin_required
def delete_status(status_id):
    status = CartridgeStatus.query.get_or_404(status_id)
    db.session.delete(status)
    event = EventLog(
        table_name='cartrg_status',
        event_type=3,  # Видалення статусу (новий тип події)
        user_updated=current_user.id
    )
    db.session.add(event)
    db.session.commit()
    flash('Статус видалено!')
    return redirect(url_for('cartridge_status'))

# Перегляд логів подій
@app.route('/event_log')
@login_required
@admin_required
def event_log():
    table_filter = request.args.get('table_filter', '')
    type_filter = request.args.get('type_filter', '')
    query = EventLog.query
    if table_filter:
        query = query.filter(EventLog.table_name == table_filter)
    if type_filter:
        query = query.filter(EventLog.event_type == int(type_filter))
    logs = query.all()
    return render_template('event_log.html', User=User, logs=logs, table_filter=table_filter, type_filter=type_filter)

#**************************робота з картриджами**************************
@app.route('/add_cartridge_event', methods=['POST'])
@login_required
def add_cartridge_event():
    serial_num = request.form.get('serial_num')
    status = request.form.get('status')
    exec_dept = request.form.get('exec_dept')
    printer = request.form.get('printer') or None
    parcel_track = request.form.get('parcel_track') or None

    # Перевірка наявності картриджа
    cartridge = Cartridges.query.filter_by(serial_num=serial_num).first()
    if not cartridge:
        return jsonify({'success': False, 'message': 'Картридж із таким серійним номером не знайдено!'})

    # Перевірка, чи вибрано відділ
    if not exec_dept:
        return jsonify({'success': False, 'message': 'Виберіть відділ!'})

    # Перевірка на дублювання статусу
    if cartridge.curr_status == int(status):
        return jsonify({'success': False, 'message': 'Цей статус уже встановлено для картриджа!'})

    # Перевірка для статусів 3 або 5, якщо is_exec == 2
    if int(status) in [3, 5]:
        dept = RefillDept.query.filter_by(id=exec_dept).first()
        if dept and dept.is_exec == 2 and cartridge.cartrg_model_id:
            # Перевірка, чи модель картриджа прив’язана до послуги
            service_mapping = CompatibleServices.query.filter_by(cartridge_model_id=cartridge.cartrg_model_id).first()
            if service_mapping:
                service = ContractsServicesBalance.query.get(service_mapping.service_id)
                if service and service.balance > 0:
                    service.balance -= 1
                    service.user_updated = current_user.id
                    service.time_updated = datetime.utcnow()
                else:
                    db.session.rollback()
                    return jsonify({'success': False, 'message': f'Недостатньо балансу для послуги {service.RefillServiceName}'})

    # Додавання події в CartridgeStatus
    new_status = CartridgeStatus(
        cartridge_id=cartridge.id,
        status=int(status),
        date_ofchange=datetime.now(),
        parcel_track=parcel_track,
        exec_dept=int(exec_dept),
        user_updated=current_user.id,
        time_updated=datetime.now()
    )
    db.session.add(new_status)

    # Оновлення запису катриджа в Cartridges
    cartridge.curr_status = int(status)
    cartridge.curr_dept = int(exec_dept)
    cartridge.curr_parcel_track = parcel_track
    cartridge.in_printer = int(printer) if printer else None
    cartridge.user_updated = current_user.id
    cartridge.time_updated = datetime.now()

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Помилка: {str(e)}'})

    return jsonify({'success': True})



@app.route('/check_cartridge', methods=['POST'])
@login_required
def check_cartridge():
    data = request.get_json()
    serial_num = data.get('serial_num')

    # Перевірка наявності картриджа
    cartridge = Cartridges.query.filter_by(serial_num=serial_num).first()
    if not cartridge:
        return jsonify({'success': False, 'message': 'Картридж не знайдено!'})

    # Отримуємо останній статус із CartridgeStatus
    latest_status = CartridgeStatus.query.filter_by(cartridge_id=cartridge.id) \
        .order_by(CartridgeStatus.date_ofchange.desc()).first()

    return jsonify({
        'success': True,
        'latest_status': {
            'status': latest_status.status if latest_status else 0,
            'exec_dept': latest_status.exec_dept if latest_status else None,
            'parcel_track': latest_status.parcel_track if latest_status else None
        }
    })

#динамічний фільтр картриджів
@app.route('/api/cartridges', methods=['GET'])
@login_required
def api_cartridges():
    """
    Повертає список картриджів із пагінацією та пошуком за всіма полями.

    Args:
        search (str): Пошуковий запит.
        page (int): Номер сторінки.

    Returns:
        JSON: Список картриджів і пагінація.
    """
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    # JOIN із CartridgeModel, CustomerEquipment, PrinterModel, RefillDept
    query = db.session.query(
        Cartridges,
        CartridgeModel.model_name,
        PrinterModel.model_name.label('printer_model_name'),
        RefillDept.deptname
    ).outerjoin(
        CartridgeModel, Cartridges.cartrg_model_id == CartridgeModel.id
    ).outerjoin(
        CustomerEquipment, Cartridges.in_printer == CustomerEquipment.id
    ).outerjoin(
        PrinterModel, CustomerEquipment.print_model == PrinterModel.id
    ).outerjoin(
        RefillDept, CustomerEquipment.print_dept == RefillDept.id
    )

    # Пошук за всіма полями
    if search:
        # Отримуємо список status_id, для яких status_name відповідає пошуковому запиту
        matching_statuses = [
            status_id for status_id, status_name in status_map.items()
            if search.lower() in status_name.lower()
        ]

        query = query.filter(
            or_(
                Cartridges.serial_num.ilike(f'%{search}%'),
                CartridgeModel.model_name.ilike(f'%{search}%'),
                PrinterModel.model_name.ilike(f'%{search}%'),
                RefillDept.deptname.ilike(f'%{search}%'),
                Cartridges.curr_status.in_(matching_statuses)
            )
        )

    # Сортування за model_name
    query = query.order_by(CartridgeModel.model_name.asc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    cartridges_data = []
    for cartridge, model_name, printer_model_name, deptname in pagination.items:
        in_printer_info = None
        if cartridge.in_printer and printer_model_name and deptname:
            in_printer_info = f"{printer_model_name} ({deptname})"

        cartridges_data.append({
            'id': cartridge.id,
            'serial_num': cartridge.serial_num,
            'cartridge_model': model_name or 'Не вказано',
            'in_printer_info': in_printer_info or 'Немає',
            'curr_status': cartridge.curr_status
        })

    pagination_data = {
        'has_prev': pagination.has_prev,
        'has_next': pagination.has_next,
        'prev_num': pagination.prev_num,
        'next_num': pagination.next_num,
        'current_page': pagination.page,
        'pages': [p if p else None for p in pagination.iter_pages(left_edge=1, left_current=2, right_current=2, right_edge=1)],
        'search': search
    }

    return jsonify({
        'cartridges': cartridges_data,
        'pagination': pagination_data
    })

#**************************end_робота з картриджами**************************
#Новий ендпоінт для принтерів
@app.route('/api/printers_by_dept/<int:dept_id>', methods=['GET'])
@login_required
def printers_by_dept(dept_id):
    printers = CustomerEquipment.query.filter_by(print_dept=dept_id).all()
    printers_data = [
        {
            'id': printer.id,
            'model_name': PrinterModel.query.get(printer.print_model).model_name
        }
        for printer in printers
    ]
    return jsonify({'printers': printers_data})

#-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-
#маршрут для отримання історії дій картриджа:
@app.route('/api/cartridge_history/<int:cartridge_id>', methods=['GET'])
@login_required
def api_cartridge_history(cartridge_id):
    history_query = db.session.query(CartridgeStatus, RefillDept.deptname, User.username)\
                              .outerjoin(RefillDept, CartridgeStatus.exec_dept == RefillDept.id)\
                              .outerjoin(User, CartridgeStatus.user_updated == User.id)\
                              .filter(CartridgeStatus.cartridge_id == cartridge_id)\
                              .order_by(CartridgeStatus.date_ofchange.desc())  # Сортуємо за датою, новіші першими
    history_data = []
    for status, dept_name, username in history_query.all():
        history_data.append({
            'date_ofchange': status.date_ofchange.isoformat(),
            'status': status.status,
            'dept_name': dept_name,
            'parcel_track': status.parcel_track,
            'user_login': username  # Змінено з user_login на username у відповідь, але в запиті правильно User.username
        })
    return jsonify({'history': history_data})

# API для асинхронного пошуку подій обробки картриджів
@app.route('/api/cartridge_status', methods=['GET'])
@login_required
def api_cartridge_status():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20

    query = db.session.query(CartridgeStatus, Cartridges.serial_num, RefillDept.deptname)\
                      .outerjoin(Cartridges, Cartridges.id == CartridgeStatus.cartridge_id)\
                      .outerjoin(RefillDept, RefillDept.id == CartridgeStatus.exec_dept)\
                      .order_by(CartridgeStatus.date_ofchange.desc())

    if search:
        query = query.filter(Cartridges.serial_num.ilike(f'%{search}%'))

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    statuses = [{
        'id': s[0].id,
        'serial_num': s[1] or 'Не вказано',
        'status': s[0].status,
        'date_ofchange': s[0].date_ofchange.isoformat(),
        'parcel_track': s[0].parcel_track,
        'deptname': s[2] or 'Не вказано'
    } for s in pagination.items]

    pagination_data = {
        'has_prev': pagination.has_prev,
        'prev_num': pagination.prev_num,
        'has_next': pagination.has_next,
        'next_num': pagination.next_num,
        'current_page': pagination.page,
        'pages': list(pagination.iter_pages()),
        'search': search
    }

    return jsonify({'statuses': statuses, 'pagination': pagination_data})


#**********************************************************
#генерація наклейки на катридж

@app.route('/api/barcode/<int:cartridge_id>', methods=['GET'])
@login_required
def generate_barcode(cartridge_id):
    cartridge = Cartridges.query.get_or_404(cartridge_id)
    serial_num = str(cartridge.serial_num)  # Перетворення в рядок

    # Транслітерація кирилиці в латинські символи
    barcode_serial = translit(serial_num, 'uk', reversed=True) if any(c.isalpha() and ord(c) > 127 for c in serial_num) else serial_num


    # Налаштування розмірів
    label_width = 80 * mm
    label_height = 25 * mm
    gap = 2 * mm

    # Створення PDF у пам'яті
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(label_width, label_height))

    # Логотип Укрпошти (чорно-білий)
    logo_path = os.path.join(app.static_folder, 'ukrposhta_logo.png')
    if os.path.exists(logo_path):
        c.drawImage(logo_path, 2 * mm, 2 * mm, width=20 * mm, height=20 * mm, preserveAspectRatio=True)

    # Генерація штрих-коду Code 128
    #barcode = Code128(serial_num, writer=ImageWriter())
    barcode = Code128(barcode_serial, writer=ImageWriter())

    barcode_path = os.path.join(app.static_folder, 'temp_barcode')
    barcode.save(barcode_path, options={"write_text": False, "module_height": 15, "module_width": 0.4})
    barcode_img_path = f"{barcode_path}.png"

    # Розміщення штрих-коду
    barcode_x = 25 * mm  # Відступ від логотипу
    barcode_y = 6 * mm  # Вертикальний відступ
    c.drawImage(barcode_img_path, barcode_x, barcode_y, width=50 * mm, height=15 * mm)

    # Текст під штрих-кодом
    c.setFont("Helvetica", 8)
    text_x = barcode_x + (50 * mm - c.stringWidth(serial_num, "Helvetica", 8)) / 2  # Центрування
    text_y = 2 * mm
    c.drawString(text_x, text_y, serial_num)

    # Завершення PDF
    c.showPage()
    c.save()

    # Очищення тимчасового файлу штрих-коду
    if os.path.exists(barcode_img_path):
        os.remove(barcode_img_path)

    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f"barcode_{serial_num}.pdf")


@app.route('/export/in_transit', methods=['GET'])
@login_required
def export_in_transit():
    # Отримуємо дані з API для "Картриджі в дорозі"
    current_date = datetime.utcnow()
    latest_status_subquery = db.session.query(CartridgeStatus.cartridge_id, func.max(CartridgeStatus.date_ofchange).label('max_date'))\
                                       .filter(CartridgeStatus.date_ofchange <= current_date)\
                                       .group_by(CartridgeStatus.cartridge_id)\
                                       .subquery()
    in_transit_query = db.session.query(Cartridges, CartridgeStatus, RefillDept.deptname, CartridgeModel.model_name) \
        .join(CartridgeStatus, Cartridges.id == CartridgeStatus.cartridge_id) \
        .outerjoin(RefillDept, CartridgeStatus.exec_dept == RefillDept.id) \
        .outerjoin(CartridgeModel, Cartridges.cartrg_model_id == CartridgeModel.id) \
        .join(latest_status_subquery,
              and_(Cartridges.id == latest_status_subquery.c.cartridge_id,
                   CartridgeStatus.date_ofchange == latest_status_subquery.c.max_date)) \
        .filter(CartridgeStatus.status == 3)
    cartridges_data = []
    for cartridge, status, dept_name, model_name in in_transit_query.all():
        cartridges_data.append({
            'id': cartridge.id,
            'serial_num': cartridge.serial_num,
            'cartridge_model': model_name or 'Не вказано',
            'date_ofchange': status.date_ofchange.strftime('%Y-%m-%d %H:%M:%S'),
            'dept_name': dept_name or 'Не вказано',
            'parcel_track': status.parcel_track or 'Не вказано'
        })

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Картриджі в дорозі"
    headers = ["ID", "Серійний номер", "Модель картриджа", "Дата зміни", "Відділ", "Трек-номер"]
    ws.append(headers)
    for cartridge in cartridges_data:
        ws.append([cartridge['id'], cartridge['serial_num'], cartridge['cartridge_model'],
                   cartridge['date_ofchange'], cartridge['dept_name'], cartridge['parcel_track']])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)  # Запас 2 символи

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))  # Межі: 10-50
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл у пам’яті
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формуємо назву файлу
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Картриджі_в_дорозі_{timestamp}.xlsx"

    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#маніпуляції з екселями
@app.route('/export/in_storage', methods=['GET'])
@login_required
def export_in_storage():
    # Запит до Cartridges із фільтром по curr_status (1 або 6) і приєднанням RefillDept

    in_storage_query = db.session.query(Cartridges, RefillDept.deptname, CartridgeModel.model_name) \
        .outerjoin(RefillDept, Cartridges.curr_dept == RefillDept.id) \
        .outerjoin(CartridgeModel, Cartridges.cartrg_model_id == CartridgeModel.id) \
        .filter(Cartridges.curr_status.in_([1, 6])) \
        .order_by(CartridgeModel.model_name.asc())
    cartridges_data = []
    for cartridge, dept_name, model_name in in_storage_query.all():
        cartridges_data.append({
            'id': cartridge.id,
            'serial_num': cartridge.serial_num,
            'cartridge_model': model_name or 'Не вказано',
            'date_ofchange': cartridge.time_updated.strftime(
                '%Y-%m-%d %H:%M:%S') if cartridge.time_updated else 'Не вказано',
            'dept_name': dept_name or 'Не вказано',
            'status': 'На зберіганні (порожній)' if cartridge.curr_status == 1 else 'На зберіганні (заправлений)'
        })

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Картриджі на зберіганні"
    headers = ["ID", "Серійний номер", "Модель картриджа", "Дата зміни", "Відділ", "Статус"]
    ws.append(headers)
    for cartridge in cartridges_data:
        ws.append([cartridge['id'], cartridge['serial_num'], cartridge['cartridge_model'],
                   cartridge['date_ofchange'], cartridge['dept_name'], cartridge['status']])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)  # Запас 2 символи

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))  # Межі: 10-50
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл у пам’яті
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формуємо назву файлу
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Картриджі_на_зберіганні_{timestamp}.xlsx"

    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/cartridge_history/<int:cartridge_id>', methods=['GET'])
@login_required
def export_cartridge_history(cartridge_id):
    # Отримуємо історію картриджа
    history_query = db.session.query(CartridgeStatus, RefillDept.deptname, User.username)\
                              .outerjoin(RefillDept, CartridgeStatus.exec_dept == RefillDept.id)\
                              .outerjoin(User, CartridgeStatus.user_updated == User.id)\
                              .filter(CartridgeStatus.cartridge_id == cartridge_id)\
                              .order_by(CartridgeStatus.date_ofchange.desc())
    history_data = []
    for status, dept_name, username in history_query.all():
        history_data.append({
            'date_ofchange': status.date_ofchange.strftime('%Y-%m-%d %H:%M:%S'),
            'status': status.status,
            'dept_name': dept_name or 'Не вказано',
            'parcel_track': status.parcel_track or 'Не вказано',
            'user_login': username or 'Не вказано'
        })

    # Отримуємо серійний номер картриджа для назви файлу з явною перевіркою типу
    cartridge = Cartridges.query.get_or_404(cartridge_id)
    serial_num = cartridge.serial_num
    if not isinstance(serial_num, str):
        serial_num = str(serial_num)
    serial_num = serial_num.replace('/', '_').replace('\\', '_')  # Замінюємо недопустимі символи

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = f"Історія_{serial_num}"
    headers = ["Дата", "Статус", "Відділ", "Трек-номер", "Оновлено користувачем"]
    ws.append(headers)
#    status_map = {
#        0: 'Не вказано',
#        1: 'На зберіганні (порожній)',
#        2: 'Відправлено в користування',
#        3: 'Відправлено на заправку',
#       4: 'Непридатний (списаний)',
#        5: 'Одноразовий (фарба у банці)',
#        6: 'На зберіганні (заправлений)'
#    }
    for event in history_data:
        ws.append([event['date_ofchange'], status_map.get(event['status'], 'Невідомий'),
                   event['dept_name'], event['parcel_track'], event['user_login']])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                # Обчислюємо довжину вмісту (перетворюємо в строку)
                value_length = len(str(cell.value))
                # Оновлюємо максимальну ширину для колонки
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)  # Додаємо запас 2 символи

    # Встановлюємо ширину колонок із обмеженнями (мін. 10, макс. 50)
    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))  # Межі ширини: від 10 до 50
        ws.column_dimensions[chr(65 + i)].width = adjusted_width  # A=65, B=66 тощо

    # Зберігаємо файл у пам’яті
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формуємо назву файлу
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Історія_{serial_num}_{timestamp}.xlsx"

    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Маршрут для експорту в Excel подій обробки картриджів
# Маршрут для експорту в Excel (змінено на /export/cartridge_status)
@app.route('/export/cartridge_events', methods=['GET'])
@login_required
def export_cartridge_events():
    search = request.args.get('search', '')

    query = db.session.query(CartridgeStatus, Cartridges.serial_num, RefillDept.deptname)\
                      .outerjoin(Cartridges, Cartridges.id == CartridgeStatus.cartridge_id)\
                      .outerjoin(RefillDept, RefillDept.id == CartridgeStatus.exec_dept)\
                      .order_by(CartridgeStatus.date_ofchange.desc())
    if search:
        query = query.filter(Cartridges.serial_num.ilike(f'%{search}%'))

    statuses = [(s[0], s[1] or 'Не вказано', s[2] or 'Не вказано') for s in query.all()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Події обробки картриджів"
    headers = ["ID", "Картридж", "Статус", "Дата зміни", "Трек-номер", "Відділ"]
    ws.append(headers)
#    status_map = {
#        0: 'Не вказано',
#        1: 'На зберіганні (порожній)',
#        2: 'Відправлено в користування',
#        3: 'Відправлено на заправку',
#        4: 'Непридатний (списаний)',
#        5: 'Одноразовий (фарба у банці)',
#        6: 'На зберіганні (заправлений)'
#    }
    for status, serial_num, deptname in statuses:
        ws.append([status.id, serial_num, status_map.get(status.status, 'Невідомий'),
                   status.date_ofchange.strftime('%Y-%m-%d %H:%M:%S'),
                   status.parcel_track or 'Немає', deptname])

    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Події_обробки_картриджів_{timestamp}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Маршрут для експорту в Excel
@app.route('/export/equipments_table', methods=['GET'])
@login_required
def export_equipments_table():
    search = request.args.get('search', '')

    query = db.session.query(CustomerEquipment, PrinterModel.model_name, RefillDept.deptname)\
                      .outerjoin(PrinterModel, PrinterModel.id == CustomerEquipment.print_model)\
                      .outerjoin(RefillDept, RefillDept.id == CustomerEquipment.print_dept)\
                      .order_by(CustomerEquipment.id)
    if search:
        query = query.filter(CustomerEquipment.inventory_num.ilike(f'%{search}%'))

    equipments = [(e[0], e[1] or 'Не вказано', e[2] or 'Не вказано') for e in query.all()]

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Список обладнання"
    headers = ["ID", "Модель", "Відділ", "Серійний номер", "Інвентарний номер"]
    ws.append(headers)
    for equip, model_name, deptname in equipments:
        ws.append([equip.id, model_name, deptname, equip.serial_num, equip.inventory_num])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Список_обладнання_{timestamp}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/cartridges_table', methods=['GET'])
@login_required
def export_cartridges_table():
    search = request.args.get('search', '')

    query = Cartridges.query
    if search:
        query = query.filter(Cartridges.serial_num.ilike(f'%{search}%'))
    cartridges = query.all()

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Список картриджів"
    headers = ["ID", "Серійний номер", "Модель картриджа", "У принтері"]
    ws.append(headers)

    for cartridge in cartridges:
        in_printer_info = "Немає"
        if cartridge.in_printer:
            equipment = CustomerEquipment.query.get(cartridge.in_printer)
            if equipment:
                printer_model = PrinterModel.query.get(equipment.print_model)
                dept = RefillDept.query.get(equipment.print_dept)
                in_printer_info = f"{printer_model.model_name} ({dept.deptname})"
        ws.append([cartridge.id, cartridge.serial_num, CartridgeModel.query.get(cartridge.cartrg_model_id).model_name or "Не вказано", in_printer_info])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл у пам’яті
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формуємо назву файлу
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Список_картриджів_{timestamp}.xlsx"

    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#*******************************************************************
#це тестовий pdf

# Реєстрація шрифту Times New Roman
pdfmetrics.registerFont(TTFont('TimesNewRoman', 'static/ttf/Times.ttf'))  # Шлях до файлу шрифту
pdfmetrics.registerFont(TTFont('TimesNewRomanBold', 'static/ttf/Timesbd.ttf'))  # Шлях до файлу шрифту

@app.route('/generate_shipping_label/<int:dept_id>', methods=['GET'])
@login_required
def generate_shipping_label(dept_id):
    # Отримуємо відділ відправника (з dept_id поточного користувача)
    sender_dept = RefillDept.query.get_or_404(current_user.dept_id)
    # Отримуємо відділ одержувача (з переданим dept_id)
    receiver_dept = RefillDept.query.get_or_404(dept_id)

    # Створюємо буфер для PDF
    buffer = BytesIO()
    textoffset = 20  # Відступ зверху

    # Ініціалізуємо PDF на розмір A4 у портретній орієнтації
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4  # 595 x 842 пунктів (портретна A4)
    half_height = height - height / 4.5

    # Заголовок "ОБЕРЕЖНО!!! НЕ КИДАТИ!!!"
    p.setFont("TimesNewRomanBold", 24)
    p.drawCentredString(width / 2, half_height + textoffset + 110, "ОБЕРЕЖНО!!!")
    p.drawCentredString(width / 2, half_height + textoffset + 80, "НЕ КИДАТИ!!!")

    # Адреса відправника (зліва, ближче до краю)
    p.setFont("TimesNewRoman", 12)
    p.drawString(20, half_height + textoffset - 10, "Відправник:")
    p.drawString(20, half_height + textoffset - 25, sender_dept.deptname or "")
    p.drawString(20, half_height + textoffset - 40, sender_dept.addr1 or "")
    p.drawString(20, half_height + textoffset - 55, sender_dept.addr2 or "")
    p.drawString(20, half_height + textoffset - 70, sender_dept.addr3 or "")
    p.drawString(20, half_height + textoffset - 85, sender_dept.addr4 or "")
    p.drawString(20, half_height + textoffset - 100, sender_dept.addr5 or "")

    horiz_offset=268
    # Адреса одержувача (справа, зсунено вліво на 1 см)
    p.drawString(width - horiz_offset, half_height + textoffset - 130, "Одержувач:")  # Змінено з -220 на -248
    p.drawString(width - horiz_offset, half_height + textoffset - 145, receiver_dept.deptname or "Не вказано")
    p.drawString(width - horiz_offset, half_height + textoffset - 160, receiver_dept.addr1 or "")
    p.drawString(width - horiz_offset, half_height + textoffset - 175, receiver_dept.addr2 or "")
    p.drawString(width - horiz_offset, half_height + textoffset - 190, receiver_dept.addr3 or "")
    p.drawString(width - horiz_offset, half_height + textoffset - 205, receiver_dept.addr4 or "")
    p.drawString(width - horiz_offset, half_height + textoffset - 220, receiver_dept.addr5 or "")

    # Номер заявки OTRS (знизу верхньої частини)
    p.drawString(20, half_height + textoffset - 240, "Номер заявки OTRS ______________________")

    # Завершуємо PDF
    p.showPage()
    p.save()

    # Повертаємо PDF як відповідь
    buffer.seek(0)
    return Response(buffer.getvalue(), mimetype='application/pdf',
                    headers={"Content-Disposition": "attachment;filename=shipping_label_"+str(dept_id)+".pdf"})

#****************** експериментально. звіт "що зроблено за день"
# Маршрут для відображення сторінки
@app.route('/report_period', methods=['GET'])
@login_required
def report_period():
    return render_template('report_period.html')


# API для отримання даних звіту
@app.route('/api/report_period', methods=['GET'])
@login_required
def api_report_period():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    # Перетворюємо рядки дат у об'єкти datetime
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        return jsonify({'error': 'Неправильний формат дати. Використовуйте YYYY-MM-DD'}), 400

    # Базовий запит до CartridgeStatus
    query = db.session.query(CartridgeStatus, Cartridges, RefillDept.deptname, User.username)\
                      .join(Cartridges, CartridgeStatus.cartridge_id == Cartridges.id)\
                      .outerjoin(RefillDept, CartridgeStatus.exec_dept == RefillDept.id)\
                      .join(User, CartridgeStatus.user_updated == User.id)\
                      .filter(CartridgeStatus.date_ofchange.between(start_date, end_date))

    # Фільтр для не-адмінів
#    if current_user.role != 'admin':  # Перевіряємо роль замість is_admin
#        query = query.filter(CartridgeStatus.user_updated == current_user.id)

    # Виконуємо запит і формуємо дані
    report_data = []
#    status_map = {
#        0: 'Не вказано',
#        1: 'На зберіганні (порожній)',
#        2: 'Відправлено в користування',
#        3: 'Відправлено на заправку',
#        4: 'Непридатний (списаний)',
#        5: 'Одноразовий (фарба у банці)',
#        6: 'На зберіганні (заправлений)'
#    }
    for status, cartridge, dept_name, username in query.order_by(CartridgeStatus.date_ofchange.desc()).all():
        report_data.append({
            'id': cartridge.id,
            'serial_num': cartridge.serial_num,
            'cartridge_model': CartridgeModel.query.get(cartridge.cartrg_model_id).model_name or 'Не вказано' if cartridge.cartrg_model_id else 'Не вказано',
            'status': status_map[status.status],
            'date_ofchange': status.date_ofchange.isoformat(),
            'dept_name': dept_name or 'Не вказано',
            'user_login': username or 'Не вказано'  # Змінено з login на username
        })

    return jsonify({'report': report_data})


@app.route('/export/report_period', methods=['GET'])
@login_required
def export_report_period():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        return "Неправильний формат дати", 400

    query = db.session.query(CartridgeStatus, Cartridges, RefillDept.deptname, User.username)\
                      .join(Cartridges, CartridgeStatus.cartridge_id == Cartridges.id)\
                      .outerjoin(RefillDept, CartridgeStatus.exec_dept == RefillDept.id)\
                      .join(User, CartridgeStatus.user_updated == User.id)\
                      .filter(CartridgeStatus.date_ofchange.between(start_date, end_date))

    if current_user.role != 'admin':  # Перевіряємо роль замість is_admin
        query = query.filter(CartridgeStatus.user_updated == current_user.id)

    report_data = []

#    status_map = {
#        0: 'Не вказано',
#        1: 'На зберіганні (порожній)',
#        2: 'Відправлено в користування',
#        3: 'Відправлено на заправку',
#        4: 'Непридатний (списаний)',
#        5: 'Одноразовий (фарба у банці)',
#        6: 'На зберіганні (заправлений)'
#    }

    for status, cartridge, dept_name, username in query.order_by(CartridgeStatus.date_ofchange.asc()).all():
        report_data.append([
            cartridge.id,
            cartridge.serial_num,
            cartridge.cartridge_model or 'Не вказано',
            status_map[status.status],
            status.date_ofchange.strftime('%Y-%m-%d %H:%M:%S'),
            dept_name or 'Не вказано',
            username or 'Не вказано'  # Змінено з login на username
        ])

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт період"
    headers = ["ID", "Серійний номер", "Модель картриджа", "Статус", "Дата зміни", "Відділ", "Користувач"]
    ws.append(headers)
    for row in report_data:
        ws.append(row)

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Звіт_період_{timestamp}.xlsx"

    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#****************** експериментально. звіт "що зроблено за день"
#****************** експериментально. масовий ввід ******************
@app.route('/mass_input')
@login_required
def mass_input():
    return render_template('mass_input.html', RefillDept=RefillDept)


@app.route('/mass_add_cartridge_events', methods=['POST'])
@login_required
def mass_add_cartridge_events():
    data = request.get_json()
    exec_dept = data.get('exec_dept')
    status = int(data.get('status'))
    serial_nums = data.get('serial_nums', [])
    parcel_track = data.get('parcel_track', '')
    printer = data.get('printer', None)

    if not exec_dept or not serial_nums:
        return jsonify({'success': False, 'message': 'Необхідно вказати відділ і хоча б один картридж!'}), 400

    # Перевірка відділу (для заправки is_exec>=1)
    is_exec_required = status in [3, 6]
    dept = RefillDept.query.filter_by(id=exec_dept).first()
    if not dept or (is_exec_required and dept.is_exec < 1):
        return jsonify({'success': False, 'message': 'Недійсний відділ для цієї операції!'}), 400

    # Оптимізоване завантаження картриджів одним запитом
    cartridges = Cartridges.query.filter(Cartridges.serial_num.in_(serial_nums)).all()
    cartridge_dict = {c.serial_num: c for c in cartridges}

    report_data = []
    invalid_cartridges = []
    status_checks = {
        2: lambda c: c.curr_status == 6,
        3: lambda c: c.curr_status == 1,
        6: lambda c: c.curr_status == 3,
        1: lambda c: c.curr_status == 2
    }

    # Перевірка для статусів 3 або 5, якщо is_exec == 2
    if status in [3, 5]:
        if dept.is_exec == 2:
            for serial_num in serial_nums:
                cartridge = cartridge_dict.get(serial_num)
                if cartridge and cartridge.cartrg_model_id:
                    service_mapping = CompatibleServices.query.filter_by(cartridge_model_id=cartridge.cartrg_model_id).first()
                    if service_mapping:
                        service = ContractsServicesBalance.query.get(service_mapping.service_id)
                        if service and service.balance > 0:
                            service.balance -= 1
                            service.user_updated = current_user.id
                            service.time_updated = datetime.utcnow()
                        else:
                            db.session.rollback()
                            return jsonify({'success': False, 'message': f'Недостатньо балансу для послуги {service.RefillServiceName}'})

    for serial_num in serial_nums:
        cartridge = cartridge_dict.get(serial_num)
        if not cartridge:
            invalid_cartridges.append(serial_num)
            continue
        if cartridge.curr_status != 0 and not status_checks[status](cartridge):
            invalid_cartridges.append(serial_num)
            continue

        # Оновлення стану картриджа
        cartridge.curr_status = status
        cartridge.curr_dept = int(exec_dept)
        cartridge.curr_parcel_track = parcel_track if status == 3 else None
        cartridge.in_printer = int(printer) if printer and status == 2 else None
        cartridge.user_updated = current_user.id
        cartridge.time_updated = datetime.now()

        # Додавання події в історію
        new_status = CartridgeStatus(
            cartridge_id=cartridge.id,
            status=status,
            date_ofchange=datetime.now(),
            exec_dept=int(exec_dept),
            parcel_track=parcel_track if status == 3 else None,
            user_updated=current_user.id,
            time_updated=datetime.now()
        )
        db.session.add(new_status)
        report_data.append({
            'serial_num': cartridge.serial_num,
            'cartridge_model': CartridgeModel.query.get(cartridge.cartrg_model_id).model_name or 'Не вказано' if cartridge.cartrg_model_id else 'Не вказано',
            'date_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    if invalid_cartridges:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Недійсні картриджі або статуси: {", ".join(invalid_cartridges)}'}), 400

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Помилка збереження: {str(e)}'}), 500

    # Генерація PDF-звіту
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setFont("TimesNewRoman", 12)
    status_titles = {
        2: "Видача заправлених картриджів",
        3: "Видача порожніх картриджів на заправку",
        6: "Прийом заправлених картриджів",
        1: "Прийом порожніх картриджів"
    }
    p.drawString(100, 800, f"Звіт: {status_titles.get(status, 'Масова операція')}")
    p.drawString(100, 780, f"Відділ: {dept.deptname}")
    p.drawString(100, 760, f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y = 740
    p.drawString(50, y, "Серійний номер")
    p.drawString(200, y, "Модель картриджа")
    p.drawString(350, y, "Дата/Час")
    y -= 20
    for item in report_data:
        if y < 50:
            p.showPage()
            p.setFont("TimesNewRoman", 12)
            y = 800
        p.drawString(50, y, item['serial_num'])
        p.drawString(200, y, item['cartridge_model'])
        p.drawString(350, y, item['date_time'])
        y -= 20
    p.showPage()
    p.save()
    buffer.seek(0)

    return Response(buffer.getvalue(), mimetype='application/pdf',
                    headers={"Content-Disposition": f"attachment;filename=mass_operation_{status}_report.pdf"})




@app.route('/api/barcodes_all', methods=['GET'])
@login_required
@admin_required
def generate_all_barcodes():
    search_query = request.args.get('search', '')
    cartridges = Cartridges.query.filter(Cartridges.serial_num.ilike(f'%{search_query}%')).all()

    if not cartridges:
        return jsonify({"error": "Немає картриджів для генерації"}), 404

    # Налаштування розмірів
    label_width = 80 * mm
    label_height = 25 * mm
    gap = 2 * mm

    # Створення PDF у пам'яті
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(label_width, label_height))

    for cartridge in cartridges:
        serial_num = str(cartridge.serial_num)
        barcode_serial = translit(serial_num, 'uk', reversed=True) if any(c.isalpha() and ord(c) > 127 for c in serial_num) else serial_num

        # Логотип Укрпошти
        logo_path = os.path.join(app.static_folder, 'ukrposhta_logo.png')
        if os.path.exists(logo_path):
            c.drawImage(logo_path, 2 * mm, 2 * mm, width=20 * mm, height=20 * mm, preserveAspectRatio=True)

        # Генерація штрих-коду Code 128
        barcode = Code128(barcode_serial, writer=ImageWriter())
        barcode_path = os.path.join(app.static_folder, f'temp_barcode_{cartridge.id}')
        barcode.save(barcode_path, options={"write_text": False, "module_height": 15, "module_width": 0.4})
        barcode_img_path = f"{barcode_path}.png"

        # Розміщення штрих-коду
        barcode_x = 25 * mm
        barcode_y = 6 * mm
        c.drawImage(barcode_img_path, barcode_x, barcode_y, width=50 * mm, height=15 * mm)

        # Текст під штрих-кодом
        c.setFont("Helvetica", 8)
        text_x = barcode_x + (50 * mm - c.stringWidth(serial_num, "Helvetica", 8)) / 2
        text_y = 2 * mm
        c.drawString(text_x, text_y, serial_num)

        # Нова сторінка для наступного картриджа
        c.showPage()

        # Очищення тимчасового файлу
        if os.path.exists(barcode_img_path):
            os.remove(barcode_img_path)

    # Завершення PDF
    c.save()
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f"all_barcodes_{datetime.now().strftime('%Y-%m-%d')}.pdf")




@app.route('/update_cartridge_barcodes', methods=['POST'])
@login_required
def update_cartridge_barcodes():
    data = request.get_json()
    cartridge_ids = data.get('cartridge_ids', [])
    barcode = data.get('barcode', '').strip()

    if not cartridge_ids or not barcode:
        return jsonify({'success': False, 'message': 'Необхідно вказати картриджі та штрих-код'}), 400

    try:
        for cartridge_id in cartridge_ids:
            # Оновлюємо parcel_track для останньої події зі статусом 3 у CartridgeStatus
            latest_status = CartridgeStatus.query.filter_by(cartridge_id=cartridge_id, status=3).order_by(CartridgeStatus.date_ofchange.desc()).first()
            if latest_status:
                latest_status.parcel_track = barcode
                latest_status.time_updated = datetime.now()
                latest_status.user_updated = current_user.id

            # Оновлюємо curr_parcel_track у таблиці Cartridges
            cartridge = Cartridges.query.get(cartridge_id)
            if cartridge:
                cartridge.curr_parcel_track = barcode
            else:
                # Якщо картридж не знайдено, логувати помилку, але не переривати цикл
                current_app.logger.error(f"Картридж з ID {cartridge_id} не знайдено в таблиці Cartridges")

        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/dashboard', methods=['GET'])
@login_required
def test_cartridges():
    return render_template('dashboard.html', RefillDept=RefillDept)


#----------------------------------------------------------------
@app.route('/api/get_cartridge_movement_all', methods=['GET'])
@login_required
def get_cartridge_movement_all():
    # Параметри запиту: рік (опціонально, за замовчуванням поточний), dept_id (опціонально)
    year = request.args.get('year', type=int, default=datetime.now().year)
    dept_id = request.args.get('dept_id', type=int, default=None)

    # Запит для підрахунку подій (статуси 2, 5) для відділів з is_exec != 1
    query = (
        db.session.query(
            RefillDept.id.label('dept_id'),
            RefillDept.deptname,
            extract('month', CartridgeStatus.date_ofchange).label('month'),
            func.count().label('event_count')
        )
        .join(RefillDept, RefillDept.id == CartridgeStatus.exec_dept)
        .filter(
            CartridgeStatus.status.in_([2, 5]),
            extract('year', CartridgeStatus.date_ofchange) == year,
            RefillDept.is_exec != 1
        )
    )

    # Додаємо фільтр за dept_id, якщо вказано
    if dept_id:
        query = query.filter(RefillDept.id == dept_id)

    query = query.group_by(
        RefillDept.id,
        RefillDept.deptname,
        extract('month', CartridgeStatus.date_ofchange)
    ).all()

    # Формування результату
    result = {}
    totals = {str(i): {'filled_sent': 0} for i in range(1, 13)}
    for dept_id, deptname, month, count in query:
        if dept_id not in result:
            result[dept_id] = {
                'deptname': deptname,
                'data': {str(i): {'filled_sent': 0} for i in range(1, 13)}
            }
        month_str = str(int(month))
        result[dept_id]['data'][month_str]['filled_sent'] = count
        totals[month_str]['filled_sent'] += count

    # Видаляємо відділи без подій
    result = {k: v for k, v in result.items() if any(
        v['data'][str(m)]['filled_sent'] > 0
        for m in range(1, 13)
    )}

    return jsonify({
        'year': year,
        'departments': result,
        'totals': totals
    })


@app.route('/cartridge_movement_all', methods=['GET'])
@login_required
def cartridge_movement_all():
    year = request.args.get('year', type=int, default=datetime.now().year)
    dept_id = request.args.get('dept_id', type=int, default=None)
    # Запит до API для отримання даних
    response = get_cartridge_movement_all().get_json()
    return render_template('cartridge_movement_all.html', RefillDept=RefillDept, year=year, dept_id=dept_id,
                           departments=response['departments'], totals=response['totals'])


@app.route('/export/cartridge_movement_all', methods=['GET'])
@login_required
def export_cartridge_movement_all():
    year = request.args.get('year', type=int, default=datetime.now().year)
    dept_id = request.args.get('dept_id', type=int, default=None)

    # Отримуємо дані з API
    response = get_cartridge_movement_all().get_json()
    departments = response['departments']
    totals = response['totals']

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = f"Заправлені картриджі, що видані на відділи {year}"

    # Заголовки: лише назви місяців
    months = ['Січ', 'Лют', 'Бер', 'Кві', 'Тра', 'Чер', 'Лип', 'Сер', 'Вер', 'Жов', 'Лис', 'Гру']
    ws.append(['Відділ'] + months)

    # Дані по відділах
    for dept_id, dept in departments.items():
        row = [dept['deptname']]
        for month in range(1, 13):
            month_str = str(month)
            row.append(dept['data'][month_str]['filled_sent'])
        ws.append(row)

    # Рядок "Всього"
    total_row = ['Всього']
    for month in range(1, 13):
        month_str = str(month)
        total_row.append(totals[month_str]['filled_sent'])
    ws.append(total_row)

    # Стилі для заголовків і рядка "Всього"
    from openpyxl.styles import Font, Alignment, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл у пам’яті
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формуємо назву файлу
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Заправлені_картриджі_що_видані_на_відділи_{year}_{timestamp}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#----------------------------------------------------------------

@app.route('/api/get_cartridge_movement_models', methods=['GET'])
@login_required
def get_cartridge_movement_models():
    # Параметри запиту: рік (опціонально, за замовчуванням поточний)
    year = request.args.get('year', type=int, default=datetime.now().year)

    # Запит для підрахунку подій (статуси 2, 5) для всіх моделей картриджів
    query = (
        db.session.query(
            CartridgeModel.id.label('model_id'),
            CartridgeModel.model_name,
            extract('month', CartridgeStatus.date_ofchange).label('month'),
            func.count().label('event_count')
        )
        .join(Cartridges, Cartridges.cartrg_model_id == CartridgeModel.id)
        .join(CartridgeStatus, CartridgeStatus.cartridge_id == Cartridges.id)
        .filter(
            CartridgeStatus.status.in_([2, 5]),
            extract('year', CartridgeStatus.date_ofchange) == year
        )
        .group_by(
            CartridgeModel.id,
            CartridgeModel.model_name,
            extract('month', CartridgeStatus.date_ofchange)
        )
    ).all()

    # Отримуємо всі моделі картриджів
    all_models = CartridgeModel.query.order_by(CartridgeModel.model_name).all()

    # Формування результату
    result = {}
    totals = {str(i): {'filled_sent': 0} for i in range(1, 13)}

    # Ініціалізація всіх моделей у result
    for model in all_models:
        result[model.id] = {
            'model_name': model.model_name,
            'data': {str(i): {'filled_sent': 0} for i in range(1, 13)}
        }

    # Заповнення даними з query
    for model_id, model_name, month, count in query:
        month_str = str(int(month))
        result[model_id]['data'][month_str]['filled_sent'] = count
        totals[month_str]['filled_sent'] += count

    return jsonify({
        'year': year,
        'models': result,
        'totals': totals
    })

@app.route('/cartridge_movement_models', methods=['GET'])
@login_required
def cartridge_movement_models():
    year = request.args.get('year', type=int, default=datetime.now().year)
    # Запит до API для отримання даних
    response = get_cartridge_movement_models().get_json()
    return render_template('cartridge_movement_models.html',
                           CartridgeModel=CartridgeModel,
                           year=year,
                           models=response['models'],
                           totals=response['totals'])

@app.route('/export/cartridge_movement_models', methods=['GET'])
@login_required
def export_cartridge_movement_models():
    year = request.args.get('year', type=int, default=datetime.now().year)

    # Отримуємо дані з API
    response = get_cartridge_movement_models().get_json()
    models = response['models']
    totals = response['totals']

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = f"Заправлені картриджі за моделями {year}"

    # Заголовки: лише назви місяців
    months = ['Січ', 'Лют', 'Бер', 'Кві', 'Тра', 'Чер', 'Лип', 'Сер', 'Вер', 'Жов', 'Лис', 'Гру']
    ws.append(['Модель картриджа'] + months)

    # Дані по моделях
    for model_id, model in models.items():
        row = [model['model_name']]
        for month in range(1, 13):
            month_str = str(month)
            row.append(model['data'][month_str]['filled_sent'])
        ws.append(row)

    # Рядок "Всього"
    total_row = ['Всього']
    for month in range(1, 13):
        month_str = str(month)
        total_row.append(totals[month_str]['filled_sent'])
    ws.append(total_row)

    # Стилі для заголовків і рядка "Всього"
    from openpyxl.styles import Font, Alignment, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл у пам’яті
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Формуємо назву файлу
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Заправлені_картриджі_за_моделями_{year}_{timestamp}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#-------------------------------------------
# Існуючий ендпоінт для отримання моделей принтерів
@app.route('/api/printer_models', methods=['GET'])
@login_required
def get_printer_models():
    printers = PrinterModel.query.all()
    return jsonify([{'id': p.id, 'model_name': p.model_name} for p in printers])


# Отримання доступних і сумісних картриджів
@app.route('/api/compatible_cartridges/<int:printer_model_id>', methods=['GET'])
@login_required
def get_compatible_cartridges(printer_model_id):
    # Перевірка існування моделі принтера
    if not PrinterModel.query.get(printer_model_id):
        return jsonify({'error': 'Printer model not found'}), 404

    compatible = CompatibleCartridges.query.filter_by(printer_model_id=printer_model_id).all()
    compatible_ids = {c.cartridge_model_id for c in compatible}
    all_cartridges = CartridgeModel.query.all()

    available = [{'id': c.id, 'model_name': c.model_name} for c in all_cartridges if c.id not in compatible_ids]
    compatible_data = []
    for c in compatible:
        cartridge = CartridgeModel.query.get(c.cartridge_model_id)
        if cartridge:
            compatible_data.append({
                'id': c.cartridge_model_id,
                'model_name': cartridge.model_name,
                'notes': c.notes
            })

    return jsonify({'available': available, 'compatible': compatible_data})


# Оновлення зв’язків сумісності
@app.route('/api/compatible_cartridges/<int:printer_model_id>', methods=['POST'])
@login_required
@admin_required
def update_compatible_cartridges(printer_model_id):
    if not PrinterModel.query.get(printer_model_id):
        return jsonify({'error': 'Printer model not found'}), 404

    data = request.json
    cartridges = data.get('cartridges', [])

    CompatibleCartridges.query.filter_by(printer_model_id=printer_model_id).delete()

    for item in cartridges:
        cartridge_model_id = item.get('cartridge_model_id')
        if not CartridgeModel.query.get(cartridge_model_id):
            db.session.rollback()
            return jsonify({'error': f'Cartridge model {cartridge_model_id} not found'}), 404

        new_link = CompatibleCartridges(
            printer_model_id=printer_model_id,
            cartridge_model_id=cartridge_model_id,
            user_updated=current_user.id,
            time_updated=datetime.utcnow(),
            notes=item.get('notes')
        )
        db.session.add(new_link)

    event = EventLog(
        table_name='compat_cartridges',
        event_type=2,
        user_updated=current_user.id
    )
    db.session.add(event)

    try:
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# Видалення конкретного зв’язку
@app.route('/api/compatible_cartridges/<int:printer_model_id>/<int:cartridge_model_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_compatible_cartridge(printer_model_id, cartridge_model_id):
    link = CompatibleCartridges.query.filter_by(
        printer_model_id=printer_model_id,
        cartridge_model_id=cartridge_model_id
    ).first()

    if not link:
        return jsonify({'error': 'Compatibility link not found'}), 404

    db.session.delete(link)

    event = EventLog(
        table_name='compat_cartridges',
        event_type=3,
        user_updated=current_user.id
    )
    db.session.add(event)

    try:
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
#------------------------------------------------------
@app.route('/cartridge_distribution_by_dept', methods=['GET'])
@login_required
def cartridge_distribution_by_dept():
    # Отримуємо рік і місяць з параметрів запиту або встановлюємо поточні
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    # Отримуємо всі моделі картриджів
    cartridge_models = CartridgeModel.query.all()

    # Запит для отримання даних: кількість картриджів за моделями та підрозділами
    results = (
        db.session.query(
            RefillDept.id.label('dept_id'),
            RefillDept.deptname.label('dept_name'),
            CartridgeModel.id.label('model_id'),
            func.count().label('count')
        )
        .join(CartridgeStatus, CartridgeStatus.exec_dept == RefillDept.id)
        .join(Cartridges, CartridgeStatus.cartridge_id == Cartridges.id)
        .join(CartridgeModel, Cartridges.cartrg_model_id == CartridgeModel.id)
        .filter(
#            CartridgeStatus.status == 2,  # Припускаємо, що status=2 означає "Видано"
            CartridgeStatus.status.in_([2, 5]),
            extract('year', CartridgeStatus.date_ofchange) == year,
            extract('month', CartridgeStatus.date_ofchange) == month
        )
        .group_by(RefillDept.id, RefillDept.deptname, CartridgeModel.id)
        .all()
    )

    # Організація даних для шаблону
    departments = {}
    totals = {}
    for result in results:
        dept_id = str(result.dept_id)
        model_id = str(result.model_id)
        if dept_id not in departments:
            departments[dept_id] = {
                'dept_name': result.dept_name,
                'data': {}
            }
        departments[dept_id]['data'][model_id] = result.count
        totals[model_id] = totals.get(model_id, 0) + result.count

    return render_template(
        'cartridge_distribution_by_dept.html',
        year=year,
        month=month,
        cartridge_models=cartridge_models,
        departments=departments,
        totals=totals
    )

@app.route('/export/cartridge_distribution_by_dept', methods=['GET'])
@login_required
def export_cartridge_distribution_by_dept():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    # Отримуємо дані (аналогічно до основного ендпоінту)
    cartridge_models = CartridgeModel.query.all()
    results = (
        db.session.query(
            RefillDept.id.label('dept_id'),
            RefillDept.deptname.label('dept_name'),
            CartridgeModel.id.label('model_id'),
            CartridgeModel.model_name.label('model_name'),
            func.count().label('count')
        )
        .join(CartridgeStatus, CartridgeStatus.exec_dept == RefillDept.id)
        .join(Cartridges, CartridgeStatus.cartridge_id == Cartridges.id)
        .join(CartridgeModel, Cartridges.cartrg_model_id == CartridgeModel.id)
        .filter(
#            CartridgeStatus.status == 2,
            CartridgeStatus.status.in_([2, 5]),
            extract('year', CartridgeStatus.date_ofchange) == year,
            extract('month', CartridgeStatus.date_ofchange) == month
        )
        .group_by(RefillDept.id, RefillDept.deptname, CartridgeModel.id, CartridgeModel.model_name)
        .all()
    )

    # Організація даних
    departments = {}
    totals = {}
    for result in results:
        dept_id = str(result.dept_id)
        model_id = str(result.model_id)
        if dept_id not in departments:
            departments[dept_id] = {
                'dept_name': result.dept_name,
                'data': {}
            }
        departments[dept_id]['data'][model_id] = result.count
        totals[model_id] = totals.get(model_id, 0) + result.count

    # Створення Excel-файлу
    wb = Workbook()
    ws = wb.active
    ws.title = f"Звіт за {month:02d}-{year}"

    # Заголовки
    headers = ['Підрозділ'] + [model.model_name for model in cartridge_models]
    ws.append(headers)

    # Дані по підрозділах
    for dept in departments.values():
        row = [dept['dept_name']] + [dept['data'].get(str(model.id), 0) for model in cartridge_models]
        ws.append(row)

    # Рядок "Всього"
    total_row = ['Всього'] + [totals.get(str(model.id), 0) for model in cartridge_models]
    ws.append(total_row)

    # Форматування
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Збереження файлу в пам’ять
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cartridge_distribution_{year}_{month:02d}.xlsx'
    )

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


# CRUD для Contracts
@app.route('/contracts')
@login_required
def contracts():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    query = db.session.query(Contracts, RefillDept.deptname)\
                      .outerjoin(RefillDept, RefillDept.id == Contracts.contractor_id)\
                      .order_by(Contracts.id)
    if search:
        query = query.filter(Contracts.contract_number.ilike(f'%{search}%'))
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    contracts = [(c[0], c[1]) for c in pagination.items]
    return render_template('contracts.html',
                           contracts=contracts,
                           search=search,
                           RefillDept=RefillDept,
                           pagination=pagination)

@app.route('/api/contracts', methods=['GET'])
@login_required
def api_contracts():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    query = db.session.query(Contracts, RefillDept.deptname)\
                      .outerjoin(RefillDept, RefillDept.id == Contracts.contractor_id)\
                      .order_by(Contracts.id)
    if search:
        query = query.filter(Contracts.contract_number.ilike(f'%{search}%'))
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    contracts = [{
        'id': c[0].id,
        'contract_number': c[0].contract_number,
        'signing_date': c[0].signing_date.strftime('%Y-%m-%d'),
        'expiry_date': c[0].expiry_date.strftime('%Y-%m-%d') if c[0].expiry_date else None,
        'contractor_id': c[0].contractor_id,
        'contractor_name': c[1] or 'Не вказано',
        'description': c[0].description,
        'status': c[0].status
    } for c in pagination.items]
    pagination_data = {
        'has_prev': pagination.has_prev,
        'prev_num': pagination.prev_num,
        'has_next': pagination.has_next,
        'next_num': pagination.next_num,
        'current_page': pagination.page,
        'pages': list(pagination.iter_pages()),
        'search': search
    }
    return jsonify({'contracts': contracts, 'pagination': pagination_data})

@app.route('/api/contract/<int:contract_id>', methods=['GET'])
@login_required
def api_contract(contract_id):
    contract = Contracts.query.get_or_404(contract_id)
    return jsonify({
        'id': contract.id,
        'contract_number': contract.contract_number,
        'signing_date': contract.signing_date.strftime('%Y-%m-%d'),
        'expiry_date': contract.expiry_date.strftime('%Y-%m-%d') if contract.expiry_date else '',
        'contractor_id': contract.contractor_id,
        'description': contract.description,
        'status': contract.status
    })

@app.route('/add_contract', methods=['POST'])
@login_required
@admin_required
def add_contract():
    data = request.get_json()
    contract_number = data.get('contract_number')
    signing_date = data.get('signing_date')
    expiry_date = data.get('expiry_date')
    contractor_id = data.get('contractor_id')
    description = data.get('description')
    status = data.get('status')
    if Contracts.query.filter_by(contract_number=contract_number).first():
        return jsonify({'success': False, 'message': 'Договір із таким номером уже існує!'}), 400
    try:
        signing_date = datetime.strptime(signing_date, '%Y-%m-%d').date()
        expiry_date = datetime.strptime(expiry_date, '%Y-%m-%d').date() if expiry_date else None
    except ValueError:
        return jsonify({'success': False, 'message': 'Невірний формат дати!'}), 400
    if not RefillDept.query.get(contractor_id):
        return jsonify({'success': False, 'message': 'Невірний підрядник!'}), 400
    if status not in ['active', 'inactive']:
        return jsonify({'success': False, 'message': 'Невірний статус!'}), 400
    contract = Contracts(
        contract_number=contract_number,
        signing_date=signing_date,
        expiry_date=expiry_date,
        contractor_id=contractor_id,
        description=description,
        status=status,
        user_updated=current_user.id,
        time_updated=datetime.utcnow()
    )
    db.session.add(contract)
    event = EventLog(
        table_name='contracts',
        event_type=1,
        user_updated=current_user.id
    )
    db.session.add(event)
    try:
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/edit_contract/<int:contract_id>', methods=['POST'])
@login_required
@admin_required
def edit_contract(contract_id):
    contract = Contracts.query.get_or_404(contract_id)
    data = request.get_json()
    contract_number = data.get('contract_number')
    signing_date = data.get('signing_date')
    expiry_date = data.get('expiry_date')
    contractor_id = data.get('contractor_id')
    description = data.get('description')
    status = data.get('status')
    if Contracts.query.filter(Contracts.contract_number == contract_number, Contracts.id != contract_id).first():
        return jsonify({'success': False, 'message': 'Договір із таким номером уже існує!'}), 400
    try:
        signing_date = datetime.strptime(signing_date, '%Y-%m-%d').date()
        expiry_date = datetime.strptime(expiry_date, '%Y-%m-%d').date() if expiry_date else None
    except ValueError:
        return jsonify({'success': False, 'message': 'Невірний формат дати!'}), 400
    if not RefillDept.query.get(contractor_id):
        return jsonify({'success': False, 'message': 'Невірний підрядник!'}), 400
    if status not in ['active', 'inactive']:
        return jsonify({'success': False, 'message': 'Невірний статус!'}), 400
    contract.contract_number = contract_number
    contract.signing_date = signing_date
    contract.expiry_date = expiry_date
    contract.contractor_id = contractor_id
    contract.description = description
    contract.status = status
    contract.user_updated = current_user.id
    contract.time_updated = datetime.utcnow()
    event = EventLog(
        table_name='contracts',
        event_type=2,
        user_updated=current_user.id
    )
    db.session.add(event)
    try:
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/delete_contract/<int:contract_id>', methods=['POST'])
@login_required
@admin_required
def delete_contract(contract_id):
    contract = Contracts.query.get_or_404(contract_id)
    db.session.delete(contract)
    event = EventLog(
        table_name='contracts',
        event_type=3,
        user_updated=current_user.id
    )
    db.session.add(event)
    db.session.commit()
    flash('Договір видалено!')
    return redirect(url_for('contracts'))

#--------------------------------
@app.route('/contract_services')
@login_required
@admin_required
def contract_services():
    return render_template('contract_services.html', Contracts = Contracts, RefillDept = RefillDept)

@app.route('/api/contract_services', methods=['GET'])
@login_required
@admin_required
def get_contract_services():
    search = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    per_page = 10

    query = ContractsServicesBalance.query.join(Contracts).join(RefillDept, Contracts.contractor_id == RefillDept.id)
    if search:
        query = query.filter(
            (Contracts.contract_number.ilike(f'%{search}%')) |
            (ContractsServicesBalance.RefillServiceName.ilike(f'%{search}%'))
        )

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    services = [{
        'id': service.id,
        'contract_id': service.contract_id,
        'contract_number': Contracts.query.get(service.contract_id).contract_number,
        'contractor_name': RefillDept.query.get(Contracts.query.get(service.contract_id).contractor_id).deptname if RefillDept.query.get(Contracts.query.get(service.contract_id).contractor_id) else 'Невідомо',
        'RefillServiceName': service.RefillServiceName,
        'service_type': service.service_type,
        'balance': service.balance,
        'initial_balance': service.initial_balance,
        'time_updated': service.time_updated.isoformat() if service.time_updated else None
    } for service in pagination.items]

    return jsonify({
        'services': services,
        'pagination': {
            'current_page': pagination.page,
            'pages': [i for i in range(1, pagination.pages + 1)],
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num,
            'search': search
        }
    })

@app.route('/api/contract_service/<int:id>', methods=['GET'])
@login_required
@admin_required
def get_contract_service(id):
    service = ContractsServicesBalance.query.get_or_404(id)
    return jsonify({
        'id': service.id,
        'contract_id': service.contract_id,
        'RefillServiceName': service.RefillServiceName,
        'service_type': service.service_type,
        'balance': service.balance
    })

@app.route('/add_contract_service', methods=['POST'])
@login_required
@admin_required
def add_contract_service():
    data = request.get_json()
    if not all(key in data for key in ['contract_id', 'RefillServiceName', 'service_type', 'balance']):
        return jsonify({'success': False, 'message': 'Відсутні обов’язкові поля'}), 400

    # Перевірка унікальності
    existing = ContractsServicesBalance.query.filter_by(
        contract_id=data['contract_id'],
        RefillServiceName=data['RefillServiceName'],
        service_type=data['service_type']
    ).first()
    if existing:
        return jsonify({'success': False, 'message': 'Така послуга вже існує для цього договору'}), 400

    # Перевірка унікальності RefillServiceName
    if ContractsServicesBalance.query.filter_by(RefillServiceName=data['RefillServiceName']).first():
        return jsonify({'success': False, 'message': 'Назва послуги вже використовується'}), 400

    service = ContractsServicesBalance(
        contract_id=data['contract_id'],
        RefillServiceName=data['RefillServiceName'],
        service_type=data['service_type'],
        balance=data['balance'],
        user_updated=current_user.id,
        time_updated=datetime.utcnow()
    )
    db.session.add(service)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/edit_contract_service/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_contract_service(id):
    service = ContractsServicesBalance.query.get_or_404(id)
    data = request.get_json()
    if not all(key in data for key in ['contract_id', 'RefillServiceName', 'service_type', 'balance']):
        return jsonify({'success': False, 'message': 'Відсутні обов’язкові поля'}), 400

    # Перевірка унікальності (окрім поточного запису)
    existing = ContractsServicesBalance.query.filter(
        ContractsServicesBalance.id != id,
        ContractsServicesBalance.contract_id == data['contract_id'],
        ContractsServicesBalance.RefillServiceName == data['RefillServiceName'],
        ContractsServicesBalance.service_type == data['service_type']
    ).first()
    if existing:
        return jsonify({'success': False, 'message': 'Така послуга вже існує для цього договору'}), 400

    # Перевірка унікальності RefillServiceName (окрім поточного запису)
    existing_name = ContractsServicesBalance.query.filter(
        ContractsServicesBalance.id != id,
        ContractsServicesBalance.RefillServiceName == data['RefillServiceName']
    ).first()
    if existing_name:
        return jsonify({'success': False, 'message': 'Назва послуги вже використовується'}), 400

    service.contract_id = data['contract_id']
    service.RefillServiceName = data['RefillServiceName']
    service.service_type = data['service_type']
    service.balance = data['balance']
    service.user_updated = current_user.id
    service.time_updated = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True})

@app.route('/delete_contract_service/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_contract_service(id):
    service = ContractsServicesBalance.query.get_or_404(id)
    db.session.delete(service)
    db.session.commit()
    return jsonify({'success': True})

#--------------------------------------------

@app.route('/api/compatible_service/<int:cartridge_model_id>', methods=['GET'])
@login_required
@admin_required
def get_compatible_service(cartridge_model_id):
    service_mapping = CompatibleServices.query.filter_by(cartridge_model_id=cartridge_model_id).first()
    if service_mapping:
        service = ContractsServicesBalance.query.get(service_mapping.service_id)
        return jsonify({
            'service_id': service_mapping.service_id,
            'contract_id': service.contract_id
        })
    return jsonify({})


@app.route('/api/compatible_service/<int:cartridge_model_id>', methods=['POST'])
@login_required
@admin_required
def save_compatible_service(cartridge_model_id):
    data = request.get_json()
    if not data.get('service_id'):
        return jsonify({'message': 'Необхідно вказати ID послуги'}), 400

    # Перевірка існування картриджа та послуги
    CartridgeModel.query.get_or_404(cartridge_model_id)
    ContractsServicesBalance.query.get_or_404(data['service_id'])

    # Перевірка існуючого запису
    existing = CompatibleServices.query.filter_by(cartridge_model_id=cartridge_model_id).first()

    if existing:
        # Якщо той самий service_id, повертаємо успіх без змін
        if existing.service_id == data['service_id']:
            return jsonify({'success': True})
        # Оновлюємо існуючий запис
        existing.service_id = data['service_id']
        existing.user_updated = current_user.id
        existing.time_updated = datetime.utcnow()
    else:
        # Створюємо новий запис
        mapping = CompatibleServices(
            cartridge_model_id=cartridge_model_id,
            service_id=data['service_id'],
            user_updated=current_user.id,
            time_updated=datetime.utcnow()
        )
        db.session.add(mapping)

    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Помилка збереження: {str(e)}'}), 500


@app.route('/api/compatible_service/<int:cartridge_model_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_compatible_service(cartridge_model_id):
    # Перевірка існування картриджа
    CartridgeModel.query.get_or_404(cartridge_model_id)

    # Видалення запису
    existing = CompatibleServices.query.filter_by(cartridge_model_id=cartridge_model_id).first()
    if existing:
        db.session.delete(existing)
        try:
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': f'Помилка видалення: {str(e)}'}), 500
    return jsonify({'success': True})


@app.route('/api/contract_services/<int:contract_id>', methods=['GET'])
@login_required
@admin_required
def get_services_by_contract(contract_id):
    services = ContractsServicesBalance.query.filter_by(contract_id=contract_id).all()
    return jsonify([{
        'id': service.id,
        'RefillServiceName': service.RefillServiceName,
        'service_type': service.service_type
    } for service in services])


@app.route('/api/decrement_service_balance/<int:service_id>', methods=['POST'])
@login_required
def decrement_service_balance(service_id):
    service = ContractsServicesBalance.query.get_or_404(service_id)
    if service.balance <= 0:
        return jsonify({'message': f'Недостатньо балансу для послуги {service.RefillServiceName}'}), 400

    try:
        service.balance -= 1
        service.user_updated = current_user.id
        service.time_updated = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'new_balance': service.balance})
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Помилка: {str(e)}'}), 500


@app.route('/inventory')
@login_required
@admin_required
def inventory():
    return render_template('inventory.html')


#=============================================API=======================================================================
#тут будуть нові маршрути, які будуть використовуватися з BluePrints
#=============================================API=======================================================================
@app.route('/api/departments', methods=['GET'])
@login_required
def api_departments():
    """
    Повертає список відділів у форматі JSON із фільтрацією за is_exec і сортуванням.

    Query-параметри:
        is_exec (int, optional): Фільтр за is_exec (0, 1 або 2). Якщо не вказано, повертаються всі відділи.
        order (str, optional): Порядок сортування ('asc' або 'desc').

    Returns:
        JSON: Список словників із полями id, deptname, dept_description, addr1-addr5.
    """
    # Отримання query-параметрів
    is_exec = request.args.get('is_exec', default=None, type=int)
    order = request.args.get('order', default='asc', type=str)

    # Валідація параметрів
    if is_exec is not None and is_exec not in {0, 1, 2}:
        return jsonify({'success': False, 'message': 'is_exec must be 0, 1, or 2'}), 400
    if order.lower() not in {'asc', 'desc'}:
        return jsonify({'success': False, 'message': 'order must be "asc" or "desc"'}), 400

    try:
        # Виклик функції getDepartmentsList
        departments = getDepartmentsList(is_exec=is_exec, order=order)
        return jsonify({'success': True, 'departments': departments})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
#=======================================================================================================================
@app.route('/api/cartridges_by_status', methods=['GET'])
@login_required
def api_cartridges_by_status():
    """
    Повертає список картриджів у форматі JSON із фільтрацією за списком статусів і сортуванням.

    Query-параметри:
        status_list (str, optional): Список статусів, розділених комами (наприклад, '1,5,6'). Якщо не вказано, повертаються всі картриджі.
        status_sort (str, optional): Порядок сортування ('asc' або 'desc'). За замовчуванням 'asc'.

    Returns:
        JSON: Список словників із полями id, serial_num, cartridge_model, status, date_ofchange, dept_name, parcel_track.
    """
    # Отримуємо параметри з запиту
    status_list_str = request.args.get('status_list')
    status_sort = request.args.get('status_sort', 'asc').lower()

    # Перевірка status_sort
    if status_sort not in ['asc', 'desc']:
        return jsonify({'error': 'status_sort must be "asc" or "desc"'}), 400

    # Обробка status_list
    status_list = None
    if status_list_str:
        try:
            # Конвертуємо status_list із рядка (наприклад, "1,5,6") у список цілих чисел
            status_list = [int(s) for s in status_list_str.split(',') if s.strip()]
            if not status_list:
                return jsonify({'error': 'status_list cannot be empty'}), 400
        except ValueError:
            return jsonify({'error': 'status_list must contain valid integers'}), 400

    try:
        # Виклик функції getCartridgesByStatus
        cartridges = getCartridgesList(status_list=status_list, status_sort=status_sort)
        return jsonify({'cartridges': cartridges})
    except Exception as e:
        return jsonify({'error': f'Error: {str(e)}'}), 500
#=======================================================================================================================
@app.route('/api/statuses', methods=['GET'])
@login_required
def get_statuses():
    """
    Отримує список статусів для випадаючих списків у модальних вікнах.

    Returns:
        JSON: Список словників із полями status_id, status_name.
    """
    statuses = getStatusList()
    return jsonify(statuses)

#=======================================================================================================================
# РОБОТА З КАРТРИДЖАМИ
#=======================================================================================================================
@app.route('/cartridges')
@login_required
@admin_required
def cartridges():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)  # Отримуємо номер сторінки з URL
    per_page = 10  # Кількість записів на сторінці (можете змінити)

    # Базовий запит із фільтром пошуку
    query = Cartridges.query.filter(Cartridges.serial_num.ilike(f'%{search}%'))
    # Додаємо пагінацію
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    cartridges = pagination.items  # Картриджі на поточній сторінці

    return render_template('cartridges.html',
                           RefillDept=RefillDept,
                           CustomerEquipment=CustomerEquipment,
                           PrinterModel=PrinterModel,
                           cartridges=cartridges,
                           CartridgeModel=CartridgeModel,
                           search=search,
                           pagination=pagination)



@app.route('/api/getCartridge', methods=['GET'])
@login_required
def get_cartridge():
    """
    Отримує дані картриджа за серійним номером.

    Args:
        serial (str): Серійний номер картриджа, переданий як query parameter.

    Returns:
        JSON: Словник із даними картриджа або повідомлення про помилку, якщо картридж не знайдено.
    """
    serial = request.args.get('serial')
    if not serial:
        return jsonify({"error": "Серійний номер не вказано"}), 400

    cartridge_data = getCartridgeData(serial)
    if not cartridge_data:
        return jsonify({"error": "Картридж не знайдено"}), 404
    return jsonify(cartridge_data), 200
#=======================================================================================================================
@app.route('/api/createCartridge', methods=['POST'])
@login_required
@admin_required
def create_cartridge():
    """
    Створює новий картридж у базі даних.

    Args:
        JSON body: Словник із полями serial_num, cartrg_model_id, in_printer, use_counter.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = createCartridgeData(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400
#=======================================================================================================================
@app.route('/api/removeCartridge', methods=['DELETE'])
@login_required
@admin_required
def remove_cartridge():
    """
    Видаляє картридж за серійним номером.

    Args:
        JSON body: Словник із полем serial_num.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    serial_num = data.get('serial_num') if data else None
    if not serial_num:
        return jsonify({"success": False, "message": "Серійний номер не вказано"}), 400

    result = removeCartridgeData(serial_num)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 404
#=======================================================================================================================
@app.route('/api/modifyCartridge', methods=['PATCH'])
@login_required
def modify_cartridge():
    """
    Оновлює дані картриджа.

    Args:
        JSON body: Словник із полями cartridge_id, serial_num, cartrg_model_id, in_printer, use_counter.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = modifyCartridgeData(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400
#=======================================================================================================================
# РОБОТА З ПРИНТЕРАМИ
#=======================================================================================================================
@app.route('/printers')
@login_required
def printers():
    """
    Рендерить сторінку зі списком принтерів.
    """
    search = request.args.get('search', '')
    return render_template('printers.html',
                           search=search,
                           RefillDept=RefillDept,
                           PrinterModel=PrinterModel)

"""
    return render_template('cartridges.html',
                           
                           CustomerEquipment=CustomerEquipment,
                           ,
                           cartridges=cartridges,
                           CartridgeModel=CartridgeModel,
                           search=search,
                           pagination=pagination)
"""

@app.route('/api/getPrinter', methods=['GET'])
@login_required
def get_printer():
    """
    Отримує дані принтера за інвентарним номером.

    Args:
        inventory_num (str): Інвентарний номер принтера.

    Returns:
        JSON: Дані принтера або помилка.
    """
    inventory_num = request.args.get('inventory_num')
    if not inventory_num:
        return jsonify({"success": False, "message": "Інвентарний номер не вказано"}), 400

    result = getPrinterData(inventory_num)
    if result["success"]:
        return jsonify(result["data"]), 200
    return jsonify({"success": False, "message": result["message"]}), 404

@app.route('/api/createPrinter', methods=['POST'])
@login_required
@admin_required
def create_printer():
    """
    Створює новий принтер.

    Args:
        JSON body: Словник із полями serial_num, inventory_num, print_model, print_dept.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = createPrinterData(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 201
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/removePrinter', methods=['DELETE'])
@login_required
@admin_required
def remove_printer():
    """
    Видаляє принтер за інвентарним номером.

    Args:
        JSON body: Словник із полем inventory_num.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    inventory_num = data.get('inventory_num') if data else None
    if not inventory_num:
        return jsonify({"success": False, "message": "Інвентарний номер не вказано"}), 400

    result = removePrinterData(inventory_num)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 404

@app.route('/api/modifyPrinter', methods=['PATCH'])
@login_required
@admin_required
def modify_printer():
    """
    Оновлює дані принтера.

    Args:
        JSON body: Словник із полями printer_id, serial_num, inventory_num, print_model, print_dept.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = modifyPrinterData(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/equipments', methods=['GET'])
@login_required
def get_equipments():
    """
    Повертає список принтерів із пагінацією та пошуком.

    Args:
        search (str): Пошуковий запит.
        page (int): Номер сторінки.

    Returns:
        JSON: Список принтерів і пагінація.
    """
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    query = db.session.query(
        CustomerEquipment,
        PrinterModel.model_name,
        RefillDept.deptname
    ).outerjoin(
        PrinterModel, PrinterModel.id == CustomerEquipment.print_model
    ).outerjoin(
        RefillDept, RefillDept.id == CustomerEquipment.print_dept
    )

    if search:
        query = query.filter(
            or_(
                CustomerEquipment.inventory_num.ilike(f'%{search}%'),
                CustomerEquipment.serial_num.ilike(f'%{search}%'),
                PrinterModel.model_name.ilike(f'%{search}%'),
                RefillDept.deptname.ilike(f'%{search}%')
            )
        )

    query = query.order_by(CustomerEquipment.id)
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    equipments = [{
        'id': equip[0].id,
        'model_name': equip[1] or 'Не вказано',
        'dept_name': equip[2] or 'Не вказано',
        'serial_num': equip[0].serial_num or 'N/A',
        'inventory_num': equip[0].inventory_num or 'N/A'
    } for equip in paginated.items]

    pagination = {
        'current_page': paginated.page,
        'total_pages': paginated.pages,
        'has_prev': paginated.has_prev,
        'has_next': paginated.has_next,
        'prev_num': paginated.prev_num,
        'next_num': paginated.next_num,
        'search': search,
        'pages': [p if p else None for p in paginated.iter_pages(left_edge=2, left_current=2, right_current=3, right_edge=2)]
    }

    return jsonify({'equipments': equipments, 'pagination': pagination})
#=======================================================================================================================
# РОБОТА З ВІДДІЛАМИ
#=======================================================================================================================
@app.route('/departments')
@login_required
def departments():
    """
    Рендерить сторінку зі списком відділів.
    """
    search = request.args.get('search', '')
    return render_template('departments.html', search=search)

@app.route('/api/getDept', methods=['GET'])
@login_required
def get_dept():
    """
    Отримує дані відділу за ID.

    Args:
        dept_id (int): ID відділу.

    Returns:
        JSON: Дані відділу або помилка.
    """
    dept_id = request.args.get('dept_id')
    if not dept_id:
        return jsonify({"success": False, "message": "ID відділу не вказано"}), 400

    result = GetDeptData(dept_id)
    if result["success"]:
        return jsonify(result["data"]), 200
    return jsonify({"success": False, "message": result["message"]}), 404

@app.route('/api/createDept', methods=['POST'])
@login_required
@admin_required
def create_dept():
    """
    Створює новий відділ.

    Args:
        JSON body: Словник із полями deptname, dept_description, addr1-addr5, is_exec.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = CreateDept(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 201
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/deleteDept', methods=['DELETE'])
@login_required
@admin_required
def delete_dept():
    """
    Видаляє відділ за ID.

    Args:
        JSON body: Словник із полем dept_id.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    dept_id = data.get('dept_id') if data else None
    if not dept_id:
        return jsonify({"success": False, "message": "ID відділу не вказано"}), 400

    result = DeleteDept(dept_id, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/modifyDept', methods=['PATCH'])
@login_required
@admin_required
def modify_dept():
    """
    Оновлює дані відділу.

    Args:
        JSON body: Словник із полями dept_id, deptname, dept_description, addr1-addr5, is_exec.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = ModifyDept(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/depts', methods=['GET'])
@login_required
def get_depts():
    """
    Повертає список відділів із пагінацією та пошуком.

    Args:
        search (str): Пошуковий запит (по deptname та is_exec).
        page (int): Номер сторінки.

    Returns:
        JSON: Список відділів і пагінація.
    """
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    query = RefillDept.query

    if search:
        # Пошук за deptname та is_exec (текст)
        is_exec_search = []
        if 'клієнт' in search.lower():
            is_exec_search.append(0)
        if 'внутрішній виконавець' in search.lower():
            is_exec_search.append(1)
        if 'зовнішній виконавець' in search.lower():
            is_exec_search.append(2)

        query = query.filter(
            or_(
                RefillDept.deptname.ilike(f'%{search}%'),
                RefillDept.is_exec.in_(is_exec_search) if is_exec_search else False
            )
        )

    query = query.order_by(RefillDept.deptname.asc())
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    depts = [{
        'id': dept.id,
        'deptname': dept.deptname,
        'dept_description': dept.dept_description,
        'addr1': dept.addr1,
        'addr2': dept.addr2,
        'addr3': dept.addr3,
        'addr4': dept.addr4,
        'addr5': dept.addr5,
        'is_exec': dept.is_exec
    } for dept in paginated.items]

    pagination = {
        'current_page': paginated.page,
        'total_pages': paginated.pages,
        'has_prev': paginated.has_prev,
        'has_next': paginated.has_next,
        'prev_num': paginated.prev_num,
        'next_num': paginated.next_num,
        'search': search,
        'pages': [p if p else None for p in paginated.iter_pages(left_edge=2, left_current=2, right_current=3, right_edge=2)]
    }

    return jsonify({'depts': depts, 'pagination': pagination})

@app.route('/export/departments_table', methods=['GET'])
@login_required
def export_departments_table():
    """
    Експортує список відділів у Excel.

    Args:
        search (str): Пошуковий запит.

    Returns:
        File: Excel-файл із даними.
    """
    search = request.args.get('search', '')

    query = RefillDept.query
    if search:
        is_exec_search = []
        if 'клієнт' in search.lower():
            is_exec_search.append(0)
        if 'внутрішній виконавець' in search.lower():
            is_exec_search.append(1)
        if 'зовнішній виконавець' in search.lower():
            is_exec_search.append(2)

        query = query.filter(
            or_(
                RefillDept.deptname.ilike(f'%{search}%'),
                RefillDept.dept_description.ilike(f'%{search}%'),
                RefillDept.addr1.ilike(f'%{search}%'),
                RefillDept.addr2.ilike(f'%{search}%'),
                RefillDept.addr3.ilike(f'%{search}%'),
                RefillDept.addr4.ilike(f'%{search}%'),
                RefillDept.addr5.ilike(f'%{search}%'),
                RefillDept.is_exec.in_(is_exec_search) if is_exec_search else False
            )
        )

    query = query.order_by(RefillDept.deptname.asc())
    depts = query.all()

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Список відділів"
    headers = ["ID", "Назва", "Опис", "Адреса 1", "Адреса 2", "Адреса 3", "Адреса 4", "Адреса 5", "Тип"]
    ws.append(headers)

    for dept in depts:
        type_label = 'Клієнт' if dept.is_exec == 0 else 'Внутрішній виконавець' if dept.is_exec == 1 else 'Зовнішній виконавець'
        ws.append([
            dept.id,
            dept.deptname,
            dept.dept_description or '',
            dept.addr1 or '',
            dept.addr2 or '',
            dept.addr3 or '',
            dept.addr4 or '',
            dept.addr5 or '',
            type_label
        ])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Список_відділів_{timestamp}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


#=======================================================================================================================
# РОБОТА З КОРИСТУВАЧАМИ
#=======================================================================================================================
@app.route('/users')
@login_required
@admin_required
def users():
    """
    Рендерить сторінку зі списком користувачів.
    """
    search = request.args.get('search', '')
    return render_template('users.html',
                           search=search,
                           RefillDept=RefillDept)

@app.route('/api/users', methods=['GET'])
@login_required
@admin_required
def get_users():
    """
    Повертає список користувачів із пагінацією та пошуком.

    Args:
        search (str): Пошуковий запит (по username та id).
        page (int): Номер сторінки.

    Returns:
        JSON: Список користувачів і пагінація.
    """
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    query = User.query

    if search:
        # Пошук за username та id
        try:
            user_id = int(search)
            query = query.filter(
                or_(
                    User.username.ilike(f'%{search}%'),
                    User.id == user_id
                )
            )
        except ValueError:
            query = query.filter(User.username.ilike(f'%{search}%'))

    query = query.order_by(User.username.asc())
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    users = [{
        'id': user.id,
        'username': user.username,
        'humanname': user.humanname,
        'dept_id': user.dept_id,
        'dept_name': RefillDept.query.get(user.dept_id).deptname if RefillDept.query.get(user.dept_id) else 'Немає',
        'role': user.role,
        'active': user.active,
        'lastlogin': user.lastlogin.isoformat() if user.lastlogin else None,
        'time_updated': user.time_updated.isoformat() if user.time_updated else None
    } for user in paginated.items]

    pagination = {
        'current_page': paginated.page,
        'total_pages': paginated.pages,
        'has_prev': paginated.has_prev,
        'has_next': paginated.has_next,
        'prev_num': paginated.prev_num,
        'next_num': paginated.next_num,
        'search': search,
        'pages': [p if p else None for p in paginated.iter_pages(left_edge=2, left_current=2, right_current=3, right_edge=2)]
    }

    return jsonify({'users': users, 'pagination': pagination})

@app.route('/api/getUser', methods=['GET'])
@login_required
@admin_required
def get_user():
    """
    Отримує дані користувача за ID.

    Args:
        user_id (int): ID користувача.

    Returns:
        JSON: Дані користувача або помилка.
    """
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({"success": False, "message": "ID користувача не вказано"}), 400

    result = GetUserData(user_id)
    if result["success"]:
        return jsonify(result["data"]), 200
    return jsonify({"success": False, "message": result["message"]}), 404

@app.route('/api/createUser', methods=['POST'])
@login_required
@admin_required
def create_user():
    """
    Створює нового користувача.

    Args:
        JSON body: Словник із полями username, password, humanname, dept_id, role, active.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = CreateUser(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 201
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/deleteUser', methods=['DELETE'])
@login_required
@admin_required
def delete_user_api():
    """
    Видаляє користувача за ID.

    Args:
        JSON body: Словник із полем user_id.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    user_id = data.get('user_id') if data else None
    if not user_id:
        return jsonify({"success": False, "message": "ID користувача не вказано"}), 400

    result = DeleteUser(user_id, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/editUser', methods=['PATCH'])
@login_required
@admin_required
def edit_user_api():
    """
    Оновлює дані користувача.

    Args:
        JSON body: Словник із полями user_id, username, password, humanname, dept_id, role, active.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = EditUser(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/export/users_table', methods=['GET'])
@login_required
@admin_required
def export_users_table():
    """
    Експортує список користувачів у Excel.

    Args:
        search (str): Пошуковий запит.

    Returns:
        File: Excel-файл із даними.
    """
    search = request.args.get('search', '')

    query = User.query
    if search:
        try:
            user_id = int(search)
            query = query.filter(
                or_(
                    User.username.ilike(f'%{search}%'),
                    User.id == user_id
                )
            )
        except ValueError:
            query = query.filter(User.username.ilike(f'%{search}%'))

    query = query.order_by(User.username.asc())
    users = query.all()

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Список користувачів"
    headers = ["ID", "Ім'я користувача", "Повне ім'я", "Відділ", "Роль", "Активний", "Останній вхід", "Оновлено"]
    ws.append(headers)

    for user in users:
        dept = RefillDept.query.get(user.dept_id)
        ws.append([
            user.id,
            user.username,
            user.humanname,
            dept.deptname if dept else 'Немає',
            'Адмін' if user.role == 'admin' else 'Користувач',
            'Так' if user.active else 'Ні',
            user.lastlogin.strftime('%Y-%m-%d %H:%M') if user.lastlogin else 'Немає',
            user.time_updated.strftime('%Y-%m-%d %H:%M') if user.time_updated else 'Немає'
        ])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Список_користувачів_{timestamp}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#=======================================================================================================================
# РОБОТА З МОДЕЛЯМИ ПРИНТЕРІВ
#=======================================================================================================================
@app.route('/printer_models')
@login_required
def printer_models():
    """
    Рендерить сторінку зі списком моделей принтерів.
    """
    search = request.args.get('search', '')
    return render_template('printer_models.html', search=search)

@app.route('/api/printermodels', methods=['GET'])
@login_required
def get_printermodels():
    """
    Повертає список моделей принтерів із пагінацією та пошуком.

    Args:
    search (str): Пошуковий запит (по model_name та id).
    page (int): Номер сторінки.

    Returns:
    JSON: Список моделей і пагінація.
    """
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    printers = PrinterModel.query

    if search:
        # Пошук за model_name та id
        try:
            model_id = int(search)
            printers = printers.filter(
                or_(
                    PrinterModel.model_name.ilike(f'%{search}%'),
                    PrinterModel.id == model_id
                )
            )
        except ValueError:
            printers = printers.filter(PrinterModel.model_name.ilike(f'%{search}%'))

    printers = printers.order_by(PrinterModel.model_name.asc())
    paginated = printers.paginate(page=page, per_page=per_page, error_out=False)

    models = [{
        'id': model.id,
        'model_name': model.model_name,
        'ink_type': model.ink_type
    } for model in paginated.items]

    pagination = {
        'current_page': paginated.page,
        'total_pages': paginated.pages,
        'has_prev': paginated.has_prev,
        'has_next': paginated.has_next,
        'prev_num': paginated.prev_num,
        'next_num': paginated.next_num,
        'search': search,
        'pages': [p if p else None for p in paginated.iter_pages(left_edge=2, left_current=2, right_current=3, right_edge=2)]
    }

    return jsonify({'models': models, 'pagination': pagination})

@app.route('/api/getPrinterModel', methods=['GET'])
@login_required
def get_printer_model():
    """
    Отримує дані моделі принтера за ID.

    Args:
        model_id (int): ID моделі.

    Returns:
        JSON: Дані моделі або помилка.
    """
    model_id = request.args.get('model_id')
    if not model_id:
        return jsonify({"success": False, "message": "ID моделі не вказано"}), 400

    result = GetPrinterModelData(model_id)
    if result["success"]:
        return jsonify(result["data"]), 200
    return jsonify({"success": False, "message": result["message"]}), 404

@app.route('/api/createPrinterModel', methods=['POST'])
@login_required
@admin_required
def create_printer_model():
    """
        Створює нову модель принтера.

    Args:
        JSON body: Словник із полями model_name, ink_type.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = CreatePrinterModel(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 201
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/deletePrinterModel', methods=['DELETE'])
@login_required
@admin_required
def delete_printer_model_api():
    """
    Видаляє модель за ID із логуванням.

    Args:
        JSON body: Словник із полем model_id.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    model_id = data.get('model_id') if data else None
    if not model_id:
        return jsonify({"success": False, "message": "ID моделі не вказано"}), 400

    result = DeletePrinterModel(model_id, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/editPrinterModel', methods=['PATCH'])
@login_required
@admin_required
def edit_printer_model_api():
    """
    Оновлює дані моделі.

    Args:
        JSON body: Словник із полями model_id, model_name, ink_type.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = EditPrinterModel(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/export/printermodels_table', methods=['GET'])
@login_required
def export_printermodels_table():
    """
    Експортує список моделей принтерів у Excel.

    Args:
        search (str): Пошуковий запит.

    Returns:
        File: Excel-файл із даними.
    """
    search = request.args.get('search', '')

    query = PrinterModel.query
    if search:
        try:
            model_id = int(search)
            query = query.filter(
                or_(
                    PrinterModel.model_name.ilike(f'%{search}%'),
                    PrinterModel.id == model_id
                )
            )
        except ValueError:
            query = query.filter(PrinterModel.model_name.ilike(f'%{search}%'))

    query = query.order_by(PrinterModel.model_name.asc())
    models = query.all()

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Список моделей принтерів"
    headers = ["ID", "Назва", "Тип чорнил"]
    ws.append(headers)

    for model in models:
        ink_type = 'Тонер' if model.ink_type == 0 else 'Рідкі чорнила' if model.ink_type == 1 else 'Стрічка'
        ws.append([
            model.id,
            model.model_name,
            ink_type
        ])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Список_моделей_принтерів_{timestamp}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#=======================================================================================================================
# РОБОТА З МОДЕЛЯМИ КАРТРИДЖІВ
#=======================================================================================================================
@app.route('/cartridge_models')
@login_required
def cartridge_models():
    """
    Рендерить сторінку зі списком моделей картриджів.
    """
    search = request.args.get('search', '')
    return render_template('cartridgemodels.html',
                           search=search,
                           PrinterModel=PrinterModel,
                           Contracts=Contracts,
                           RefillDept=RefillDept)

@app.route('/api/cartridgemodels', methods=['GET'])
@login_required
def get_cartridgemodels():
    """
    Повертає список моделей картриджів із пагінацією та пошуком.

    Args:
        search (str): Пошуковий запит (по model_name та id).
        page (int): Номер сторінки.

    Returns:
        JSON: Список моделей і пагінація.
    """
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    query = CartridgeModel.query

    if search:
        # Пошук за model_name та id
        try:
            model_id = int(search)
            query = query.filter(
                or_(
                    CartridgeModel.model_name.ilike(f'%{search}%'),
                    CartridgeModel.id == model_id
                )
            )
        except ValueError:
            query = query.filter(CartridgeModel.model_name.ilike(f'%{search}%'))

    query = query.order_by(CartridgeModel.model_name.asc())
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    models = []
    for model in paginated.items:
        has_service = CompatibleServices.query.filter_by(cartridge_model_id=model.id).first() is not None
        printer = PrinterModel.query.get(model.printer_model_id) if model.printer_model_id else None
        models.append({
            'id': model.id,
            'model_name': model.model_name,
            'model_type': model.model_type,
            'printer_model_id': model.printer_model_id,
            'printer_model_name': printer.model_name if printer else None,
            'has_service': has_service
        })

    pagination = {
        'current_page': paginated.page,
        'total_pages': paginated.pages,
        'has_prev': paginated.has_prev,
        'has_next': paginated.has_next,
        'prev_num': paginated.prev_num,
        'next_num': paginated.next_num,
        'search': search,
        'pages': [p if p else None for p in paginated.iter_pages(left_edge=2, left_current=2, right_current=3, right_edge=2)]
    }

    return jsonify({'models': models, 'pagination': pagination})

@app.route('/api/getCartridgeModel', methods=['GET'])
@login_required
def get_cartridge_model():
    """
    Отримує дані моделі картриджа за ID.

    Args:
        model_id (int): ID моделі.

    Returns:
        JSON: Дані моделі або помилка.
    """
    model_id = request.args.get('model_id')
    if not model_id:
        return jsonify({"success": False, "message": "ID моделі не вказано"}), 400

    result = GetCartridgeModelData(model_id)
    if result["success"]:
        return jsonify(result["data"]), 200
    return jsonify({"success": False, "message": result["message"]}), 404

@app.route('/api/createCartridgeModel', methods=['POST'])
@login_required
@admin_required
def create_cartridge_model():
    """
    Створює нову модель картриджа.

    Args:
        JSON body: Словник із полями model_name, model_type, printer_model_id (опціонально).

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = CreateCartridgeModel(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 201
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/deleteCartridgeModel', methods=['DELETE'])
@login_required
@admin_required
def delete_cartridge_model_api():
    """
    Видаляє модель за ID із логуванням.

    Args:
        JSON body: Словник із полем model_id.

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    model_id = data.get('model_id') if data else None
    if not model_id:
        return jsonify({"success": False, "message": "ID моделі не вказано"}), 400

    result = DeleteCartridgeModel(model_id, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/api/editCartridgeModel', methods=['PATCH'])
@login_required
@admin_required
def edit_cartridge_model_api():
    """
    Оновлює дані моделі.

    Args:
        JSON body: Словник із полями model_id, model_name, model_type, printer_model_id (опціонально).

    Returns:
        JSON: Результат операції {"success": bool, "message": str}.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "Дані не надіслано"}), 400

    result = EditCartridgeModel(data, current_user.id)
    if result["success"]:
        return jsonify({"success": True, "message": result["message"]}), 200
    return jsonify({"success": False, "message": result["message"]}), 400

@app.route('/export/cartridgemodels_table', methods=['GET'])
@login_required
def export_cartridgemodels_table():
    """
    Експортує список моделей картриджів у Excel.

    Args:
        search (str): Пошуковий запит.

    Returns:
        File: Excel-файл із даними.
    """
    search = request.args.get('search', '')

    query = CartridgeModel.query
    if search:
        try:
            model_id = int(search)
            query = query.filter(
                or_(
                    CartridgeModel.model_name.ilike(f'%{search}%'),
                    CartridgeModel.id == model_id
                )
            )
        except ValueError:
            query = query.filter(CartridgeModel.model_name.ilike(f'%{search}%'))

    query = query.order_by(CartridgeModel.model_name.asc())
    models = query.all()

    # Створюємо Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Список моделей картриджів"
    headers = ["ID", "Назва моделі", "Тип", "Модель принтера"]
    ws.append(headers)

    for model in models:
        model_type = {0: 'Тонер + барабан', 1: 'Тільки тонер', 2: 'Тільки барабан', 3: 'Стрічка', 4: 'Чорнила'}.get(model.model_type, 'Невідомо')
        printer = PrinterModel.query.get(model.printer_model_id) if model.printer_model_id else None
        ws.append([
            model.id,
            model.model_name,
            model_type,
            printer.model_name if printer else 'Не вказано'
        ])

    # Налаштування ширини колонок
    column_widths = {}
    for row in ws.rows:
        for i, cell in enumerate(row):
            if cell.value:
                value_length = len(str(cell.value))
                column_widths[i] = max(column_widths.get(i, 10), value_length + 2)

    for i, width in column_widths.items():
        adjusted_width = max(10, min(width, 50))
        ws.column_dimensions[chr(65 + i)].width = adjusted_width

    # Зберігаємо файл
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Список_моделей_картриджів_{timestamp}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')





if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)